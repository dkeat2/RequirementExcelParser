import openpyxl
import re
import sys


class Requirement(object):
    def __init__(self, req_id):
        self.req_id = req_id
        self.req_text = []
        self.parents = []
        self.children = []

        self.multiple_parents = False
        self.multiple_texts = False


# function to map column names to their index within the worksheet
def get_col_name_lookup(sheet):
    name2index = { }
    i = 0
    for col in sheet.iter_cols(1, sheet.max_column):
        name2index[col[0].value] = i
        i += 1
    return name2index


class RequirementExtractor(object):
    def __init__(self):
        self.REQUIREMENTS = { }

    def parse_workbook(self, workbook):
        for sheet in workbook.worksheets:
            col_name_lookup = get_col_name_lookup(sheet)  # get a mapping between col name and index
            sheetname = sheet.title
            n_rows = sheet.max_row
            for row_cells in sheet.iter_rows(min_row=2, max_row=n_rows):

                req = None

                # get the requirement ID
                req_id = row_cells[col_name_lookup['Req ID']].value

                if req_id and (req_id not in self.REQUIREMENTS):
                    self.REQUIREMENTS[req_id] = Requirement(req_id)
                    req = self.REQUIREMENTS[req_id]
                elif req_id and (req_id in self.REQUIREMENTS):
                    req = self.REQUIREMENTS[req_id]

                # try to get the requirement text (MCs have none)
                req_text = None
                if ('%s Requirement Text' % sheetname) in col_name_lookup:
                    req_text = row_cells[col_name_lookup['%s Requirement Text' % sheetname]].value
                    req.req_text.append(req_text)

                # try to get the requirement parent text (MCs have none)
                req_parent_text = None
                if ('Parent') in col_name_lookup:
                    req_parent_text = row_cells[col_name_lookup['Parent']].value
                    self.process_parents(req, req_parent_text)

                # try to get the functional children text
                req_func_children_text = None
                if ('Functional Child') in col_name_lookup:
                    req_func_children_text = row_cells[col_name_lookup['Functional Child']].value
                    self.process_children(req, req_func_children_text)
        self.postprocess()
        self.export()

    def process_parents(self, req, req_parent_text):
        parent_id_match = re.match(r'(?sm).*\n(\w+\-\w+\-\d+)\:(.*)', req_parent_text)
        parent_id = None
        if parent_id_match:
            parent_id = parent_id_match.group(1)

        if parent_id and (parent_id not in self.REQUIREMENTS):
            parent_req = Requirement(parent_id)
            parent_req.req_text.append(parent_id_match.group(2).strip())
            parent_req.children.append(req.req_id)
            req.parents.append(parent_id)
            self.REQUIREMENTS[parent_id] = parent_req
        elif parent_id and (parent_id in self.REQUIREMENTS):
            parent_req = self.REQUIREMENTS[parent_id]
            parent_req.req_text.append(parent_id_match.group(2).strip())
            parent_req.children.append(req.req_id)
            req.parents.append(parent_id)

    def process_children(self, req, req_children_text):
        if not req_children_text:
            return
        children = req_children_text.split("\n")
        for c in children:
            child_id_match = parent_id_match = re.match(r'(\w+\-\w+\-\d+)\:(.*)', c)
            if child_id_match:
                child_id = child_id_match.group(1)
                child_text = child_id_match.group(2).strip()

                if child_id and (child_id not in self.REQUIREMENTS):
                    child_req = Requirement(child_id)
                    child_req.req_text.append(child_text)
                    req.children.append(child_id)
                    child_req.parents.append(req.req_id)
                    self.REQUIREMENTS[child_id] = child_req
                elif child_id and (child_id in self.REQUIREMENTS):
                    child_req = self.REQUIREMENTS[child_id]
                    child_req.req_text.append(child_text)
                    req.children.append(child_id)
                    child_req.parents.append(req.req_id)

    def postprocess(self):
        for r in self.REQUIREMENTS.values():
            r.parents = list(set(r.parents))
            r.children = list(set(r.children))
            r.req_text = list(set(r.req_text))

            r.multiple_parents = len(r.parents) > 1
            r.multiple_texts = len(r.req_text) > 1
            if len(r.req_text) == 1:
                r.req_text = r.req_text[0]
        multiple_parent_reqs = [x for x in self.REQUIREMENTS.values() if x.multiple_parents]
        if len(multiple_parent_reqs) > 0:
            print("The following requirements have multiple parents:")
            for r in multiple_parent_reqs:
                print("\t%s" % r.req_id)

        multiple_text_reqs = [x for x in self.REQUIREMENTS.values() if x.multiple_texts]
        if len(multiple_text_reqs) > 0:
            print("The following requirements have multiple texts. Please select which text to use.")
            for r in multiple_text_reqs:
                print("\n\t%s:" % r.req_id)
                for i, t in enumerate(r.req_text):
                    print("\t\t%s: %s" % (i + 1, t))
                ind = input("\n\t\tSelected Text Number: ")
                while int(ind) > len(r.req_text) or int(ind) < 1:
                    print("\t\tPlease select a valid choice.")
                    ind = input("\n\t\tSelected Text Number: ")
                r.req_text = r.req_text[int(ind) - 1]

    def export(self):
        fname = input("\n Chose a filename for export: ")
        with open(fname, 'w') as f:
            f.write("id,text,parents,children\n")
            for r in self.REQUIREMENTS.values():
                f.write("%s,%s,%s,%s\n" % (r.req_id, r.req_text, ";".join(r.parents), ";".join(r.children)))
            f.close()


if __name__ == "__main__":
    workbook_name = sys.argv[1]
    E = RequirementExtractor()
    E.parse_workbook(openpyxl.load_workbook(workbook_name))
